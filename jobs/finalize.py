import argparse
import configparser
import datetime
import sys
from os import environ
from os import path
from pathlib import Path
from shutil import copytree
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from styleframe import StyleFrame
import matplotlib.pyplot as plt
import numpy as np
from scipy.interpolate import make_interp_spline, BSpline

ILURL_HOME = environ['ILURL_HOME']
CONFIG_PATH = ILURL_HOME + '/data/plots'


def get_metric_data(dst_folder, metric_name):
    data = pd.read_csv(dst_folder + "/test/" + metric_name + "_stats.csv", names=["name", "data"])
    mean = data[data["name"] == "mean"]["data"].values[0]
    std = data[data["name"] == "std"]["data"].values[0]
    return '{0:.3f} ± {1:.3f}'.format(mean, std)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        header = False
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, header=None, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def get_arguments():
    parser = argparse.ArgumentParser(
        description="""
            This scripts runs recursively a rollout for every checkpoint stored
            on the experiment path. If test is set to True only the last checkpoints
            will be used.
        """
    )
    parser.add_argument('experiment_dir', type=str, nargs='?',
                        help='''A directory which it\'s subdirectories are train runs.''')

    parsed = parser.parse_args()
    sys.argv = [sys.argv[0]]

    return parsed


def smooth_data(data):
    x = np.array(list(range(len(data))))
    y = np.array(list(data))

    xnew = np.linspace(x.min(), x.max(), 1000)

    spl = make_interp_spline(x, y, k=7)
    y_smooth = spl(xnew)
    return xnew, y_smooth

def create_loss_graph(data, dst):
    xnew, y_smooth = smooth_data(data)
    fig, ax = plt.subplots(figsize=(24, 6))
    plt.plot(xnew, y_smooth)
    plt.xlabel('steps')
    plt.title('loss')
    plt.ylim([0, max(y_smooth) * 1.1])
    plt.grid()
    plt.plot(data, alpha=0.1)
    plt.savefig(os.path.join(dst, 'loss.png'))
    return os.path.join(dst, 'loss.png')


def finalize(experiment_dir=None, time=0, filename="Results"):
    if not experiment_dir:
        args = get_arguments()
        batch_path = Path(args.experiment_dir)
    else:
        batch_path = Path(experiment_dir)

    experiment_name = str(batch_path).split("/")[-1]
    dst_folder = CONFIG_PATH + "/" + experiment_name

    if not path.exists(dst_folder):
        copytree(str(batch_path) + "/plots", dst_folder)

    lossdata =  pd.read_csv(list(Path(batch_path).rglob('logs/*/*learning.csv'))[0])
    lossdata = lossdata['loss'].map(lambda x: float(x.split("tf.Tensor(")[1].split(", shape=")[0]))[15:]

    # create smooth line chart
    dst = create_loss_graph(lossdata, dst_folder)

    train_config_path = list(Path(batch_path).rglob('train.config'))[0]
    train_config = configparser.ConfigParser()
    train_config.read(train_config_path)

    network = train_config['train_args']['network']
    demand_type = train_config['train_args']['demand_type']
    demand_mode = train_config['train_args']['demand_mode']
    exp_time = train_config['train_args']['experiment_time']
    tls_type = train_config['train_args']['tls_type']
    agent_type = train_config['agent_type']['agent_type']
    features = train_config['mdp_args']['features'].replace("\'", "")
    reward = train_config['mdp_args']['reward'].replace("\'", "")
    travel_time = get_metric_data(dst_folder, "travel_time")
    waiting_time = get_metric_data(dst_folder, "waiting_time")
    speed = get_metric_data(dst_folder, "speed")
    stops = get_metric_data(dst_folder, "stops")
    replay_min = train_config['dqn_args']['min_replay_size']
    replay_max = train_config['dqn_args']['max_replay_size']
    eps_time = train_config['dqn_args']['epsilon_schedule_timesteps']
    test_actions_per_intersection = '=HYPERLINK("{}", "{}")'.format(
        "data/plots/" + experiment_name + "/test/actions_per_intersection.png", "Image")
    test_rewards_per_intersection = '=HYPERLINK("{}", "{}")'.format(
        "data/plots/" + experiment_name + "/test/rewards_per_intersection.png", "Image")
    train_actions_per_intersection = '=HYPERLINK("{}", "{}")'.format(
        "data/plots/" + experiment_name + "/train/actions_per_intersection.png", "Image")
    train_rewards_per_intersection = '=HYPERLINK("{}", "{}")'.format(
        "data/plots/" + experiment_name + "/train/rewards_per_intersection.png", "Image")
    loss = '=HYPERLINK("{}", "{}")'.format(
        "data/plots/" + experiment_name + "/loss.png", "Image")
    rewards = '=HYPERLINK("{}", "{}")'.format(
        "data/plots/train" + experiment_name + "/rewards.png", "Image")
    network_size = train_config['dqn_args']['torso_layers'] + train_config['dqn_args']['head_layers']

    data = pd.DataFrame(data=
                        {"Filename": experiment_name, "Network": network, "Demand Type": demand_type,
                         "Demand mode": demand_mode,
                         "Exp. Time": exp_time,
                         "Exp. Time2": int(exp_time) / 60,
                         "TLS Type": tls_type, "Agent Type": agent_type, "Features": features, "Reward": reward,
                         "Travel Time": travel_time,
                         "Waiting Time": waiting_time,
                         "Speed": speed, "Stops": stops,
                         "Epsilon Time": eps_time,
                         "Min Replay": replay_min,
                         "Max Replay": replay_max,
                         "Time": str(datetime.timedelta(seconds=int(time))),
                         "Train: Actions Per Intersection": train_actions_per_intersection,
                         "Train: Rewards Per Intersection": train_rewards_per_intersection,
                         "Test: Actions Per Intersection": test_actions_per_intersection,
                         "Test: Rewards Per Intersection": test_rewards_per_intersection,
                         "Loss": loss,
                         "Rewards": rewards,
                         "Network Size": network_size,
                         }, index=[0])
    append_df_to_excel(ILURL_HOME + "/" + filename + ".xlsx", data, index=False)


def redo_file(filename="Results"):
    exps = os.listdir(os.path.join(ILURL_HOME, 'data/emissions/'))
    exps.sort()
    for exp in exps:
        finalize(os.path.join(ILURL_HOME, 'data/emissions/' + exp), filename=filename)



if __name__ == '__main__':
    finalize()
    # redo_file()